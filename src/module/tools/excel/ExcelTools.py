from src.module.config import *
from openpyxl import *


class ExcelTools:

    def __init__(self, name, sheetDefault=0):
        self.path = GF.excel_path()
        self.name = None
        self.workFile = None
        self.sheet = None
        self._sheetIndex = 0
        self.reload(name, sheetDefault)


    def reload(self, name, sheetIndex=0):
        """
        重新加载excel
        :param name:
        :param sheetIndex:
        :return:
        """
        self.name = name
        self.workFile = load_workbook(self.path + self.name)
        # 通过索引获取表格
        sheets = self.get_sheet_names()
        self._sheetIndex = sheetIndex
        if len(sheets) > 0:
            self.sheet = self.workFile[sheets[self._sheetIndex]]

        assert self.sheet is not None, 'Excel Sheet为空'

    def flash(self):
        """
        刷新excel
        :return:
        """
        self.reload(self.name, self._sheetIndex)


    def get_sheet_names(self):
        """
        获取表格名字
        :return:
        """
        return self.workFile.get_sheet_names()


    def get_sheet_shape(self):
        """
        获取表格形状
        :return:(行,列)
        """
        return self.sheet.max_row, self.sheet.max_column

    def get_cell(self, row_index, col_index):
        """
        获取点数据
        :param row_index:
        :param col_index:
        :return:
        """
        rg1 = self.__makeSlice__(slice(row_index, row_index - 1, 1), self.sheet.max_row)
        rg2 = self.__makeSlice__(slice(col_index, col_index - 1, 1), self.sheet.max_column)
        retLst = [item for item in
                  self.sheet.iter_rows(min_row=rg1.start, max_row=rg1.stop, min_col=rg2.start, max_col=rg2.stop)]
        return retLst[0][0]

    def __getitem__(self, index: slice or tuple or int) -> list or str:
        """
        重载操作符,实现批量获取
        如:excel[2:5:2,2:5:1],excel[2:5:2],excel[:,2:5:1]
        :param index:
        :return:
        """
        rg1, rg2 = None, None

        # 1表示返回list,0表示返回str
        rg1_type, rg2_type = 1, 1
        if isinstance(index, int):
            rg1, rg1_type = self.__checkSlice__(index)
            rg1 = self.__makeSlice__(rg1, self.sheet.max_row)
            rg2 = slice(0, self.sheet.max_column, 1)

        elif isinstance(index, slice):
            rg1 = self.__makeSlice__(index, self.sheet.max_row)
            rg2 = slice(0, self.sheet.max_column, 1)

        elif isinstance(index, tuple):

            rg1, rg1_type = self.__checkSlice__(index[0])
            rg1 = self.__makeSlice__(rg1, self.sheet.max_row)

            rg2, rg2_type = self.__checkSlice__(index[1])
            rg2 = self.__makeSlice__(rg2, self.sheet.max_column)

        retLst = [item for item in
                  self.sheet.iter_rows(min_row=rg1.start, max_row=rg1.stop, min_col=rg2.start, max_col=rg2.stop,
                                       values_only=True)]
        if rg1.step != 1:
            retLst = retLst[::rg1.step]

        if rg2.step != 1:
            retLst = [item[::rg2.step] for item in retLst]

        if rg2_type == 0:
            retLst = [item[0] for item in retLst]

        if rg1_type == 0:
            retLst = retLst[0]

        return retLst

    def __checkSlice__(self, index: slice or int):
        assert isinstance(index, (slice, int)), "index错误的类型"
        if isinstance(index, int):
            # 返回数字
            return slice(index, index, 1), 0
        else:
            return index, 1

    # 切片编辑
    def __makeSlice__(self, index: slice, length):
        start = index.start if index.start is not None else 1
        stop = index.stop if index.stop is not None else length
        step = index.step if index.step is not None else 1

        assert start > 0, "输入初始坐标错误,请大于1"
        assert stop > 0, "输出初始坐标错误,请大于1"
        assert stop <= length, "长度超出限制,输入:" + str(stop) + "目标:" + str(length)
        assert step > 0, "步长错误"
        assert start <= stop, "区间错误"

        return slice(start, stop, step)
